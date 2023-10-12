import openpyxl
import re
from datetime import date



class ExcelService:
    def __init__(self,category_name,category_id,cards) -> None:
        self.category_name=category_name
        self.category_id=category_id
        self.product_cards=cards
        self.run_date = date.today()
        self.ranges = [
                (41, 54),
                (55, 69),
                (70, 75),
                (76, 80),
                (81, 89),
                (90, 109),
                (110, 119),
                (120, 139),
                (140, 180),
                (180, 199),
                (200, 230)
            ]

    def save_data_to_excel(self):
         # Создаем новый Excel-документ
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Записываем self.category_name в первую ячейку
        worksheet.cell(row=1, column=1, value=self.category_name)
        lak_data = []  # Создаем пустой список наших товаров
        sellers=['1AK']
        second_headers=['Артикул','Наименование','Бренд','Цена','Цена со скидкой']
        line_data=[]

    # Функция для сегментации данных
        for data in self.product_cards:
            if '1AK.RU' in data['Продавец']:
                lak_data.append(data)  # Добавляем текущий элемент в новый список
                self.product_cards.remove(data)  # Удаляем текущий элемент из исходного списка

        if self.category_name=="cat=128636":
            
            for data in lak_data:
                
                for element in self.product_cards:
                    
                    for start, end in self.ranges:
                        
                        if start<=int(re.search(r'\d+',element['Емкость']).group)<=end and start<=int(re.search(r'\d+',data['Емкость']).group)<=end:
                            if int(data['Пусковой ток'])-100<=int(element['Пусковой ток'])<=int(data['Пусковой ток'])-100:
                                if data['Полярность'].lower()in element['Полярность'].lower():
                                    sellers.append(element['Продавец'])
                                    second_headers.append('Бренд','Ссылка','Цена','Цена со скидкой')
                                    line_data.append(data['Артикул'],data['Наименование'],data['Бренд'],data['Цена'],data['Цена со скидкой'],element['Бренд'],element['Ссылка'],element['Цена'],element['Цена со скидкой'])
                
                for seller in sellers:
                    # Используем merge_cells для объединения ячеек в 1 строку и 5 столбцов
                    worksheet.merge_cells(start_row=2, start_column=sellers.index(seller) * 5 + 1, end_row=2, end_column=(sellers.index(seller) + 1) * 5)
                    # Записываем значение в первую ячейку объединенной области
                    worksheet.cell(row=2, column=sellers.index(seller) * 5 + 1, value=seller)
                
                worksheet.append(second_headers)
                worksheet.append(line_data)
                sellers=['1AK']
                line_data=[]
                second_headers=['Артикул','Наименование','Бренд','Цена','Цена со скидкой']
                workbook.save("wildberries_",
                              f"{self.run_date.strftime('%Y-%m-%d')}xlsx")
                
        elif self.category_name=="cat=128306":
           for data in lak_data:
                for element in self.product_cards:
                    pass
                sellers=[]
                counting=[]       

        elif self.category_name=="cat=59132":
            for data in lak_data:
                for element in self.product_cards:
                    pass
                sellers=[]
                counting=[]  

        elif self.category_name=="cat=128697":
            for data in lak_data:
                for element in self.product_cards:
                    pass
                sellers=[]
                counting=[]            

        elif self.category_name=="subject=3906":
            for data in lak_data:
                for element in self.product_cards:
                    pass
                sellers=[]
                counting=[]             

        return lak_data